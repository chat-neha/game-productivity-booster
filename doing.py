#import module
from tkinter import *
from tkinter import messagebox
import importlib
import coins as cn

from openpyxl import Workbook
from openpyxl import load_workbook 

wb = load_workbook("Tasks.xlsx")
wsheet = wb.active
                   
#create mainloop 
def doing_main():
    global ws
    ws =Tk()
    ws.geometry('800x400+500+200')
    ws.title('Your Tasklist')
    ws.config(bg='#8ebebd')
    ws.resizable(width=False,height=False)

    frame=Frame(ws)
    frame.pack(pady=10)
    global lb
    lb=Listbox(
        frame,#putting the listbox on the frame  
        width=25,
        height=8,
        font=('Courier',16, 'bold'),
        bd=0,
        fg='#464646',
        highlightthickness=10,
        selectbackground='#000000',
        activestyle="none",
    )
    lb.pack(side=LEFT, fill=BOTH)

    for row in wsheet.iter_rows(values_only=True):
        task = row[0]
        lb.insert(END, task)

    sb=Scrollbar(frame, activebackground = '#5da2a1')
    sb.pack(side=RIGHT, fill=Y)

    lb.config(yscrollcommand=sb.set)
    sb.config(command=lb.yview)
    global my_entry
    my_entry=Entry(
        ws,
        font=('Courier',24, 'bold')
        )

    my_entry.pack(pady=20)

    button_frame=Frame(ws)
    button_frame.pack(pady=20)

    addTask_btn=Button(
        button_frame,
        text='Add Task',
        font=('Courier', 14, 'bold'),
        bg='#c5f776',
        padx=20,
        pady=10,
        command=newTask
    )

    addTask_btn.pack(fill=BOTH, expand=True, side=LEFT)

    delTask_btn = Button(
        button_frame,
        text='Delete Task',
        font=('Courier', 14, 'bold'),
        bg='#ff8b61',
        padx=20,
        pady=10,
        command=deleteTask
    )
        
    delTask_btn.pack(fill=BOTH,expand=True,side=LEFT)

    return_btn = Button(
        button_frame,
        text = 'Return to Menu',
        font = ('Courier', 14, 'bold'),
        bg = '#ff8b61',
        padx = 20,
        pady = 10,
        command = goBack
        )

    return_btn.pack(fill=BOTH,expand=True,side=LEFT)





    ws.mainloop() 


def newTask():
    task=my_entry.get()#storing the task entered by the user in taask variable
    if task !="":
        lb.insert(END, task)#indicates every new task will be added at the end of the listbox
        wsheet.append([task,])
        my_entry.delete(0, "end")#to erase the new task from the entry box
        wb.save("Tasks.xlsx")
    else:
        messagebox.showwarning("Warning","Please enter some task")
        
def deleteTask():
    selected_task_index = lb.curselection()
    
    if selected_task_index:
        
        task_index = selected_task_index[0] #index of selected task in listbox
        
        lb.delete(task_index) # Delete the selected task from the Listbox

        cn.coin_increase(100)
        
##        wb = load_workbook("Tasks.xlsx")
##        wsheet = wb.active
        
        rowcount=wsheet.max_row
        
        for row in range (rowcount):
            if row == task_index:
                wsheet.delete_rows(row+1) #because .xlsx file mein index 1 zyaada hota hai

        
        wb.save("Tasks.xlsx")

        
        
    else:
        messagebox.showwarning("Warning", "Please select a task to delete")

def goBack():
    ws.destroy()
    main = importlib.import_module('menu')
    main.main_menu()
    
